import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .utils import get_unit
from .constants import *

# ── Colours ───────────────────────────────────────────────────────────────
NAVY        = "1F3864"
WHITE       = "FFFFFF"
ORANGE      = "F4B942"
LIGHT_BLUE  = "BDD7EE"
MID_BLUE    = "2E75B6"
GREY_ROW    = "F2F2F2"
HEADER_BLUE = "4472C4"

thin  = Side(style="thin",   color="CCCCCC")
thick = Side(style="medium", color="000000")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(colour): return PatternFill("solid", fgColor=colour)

def style_cell(cell, bold=False, fg=None, bg=None, align="center",
               size=10, wrap=False, italic=False, valign="center"):
    cell.font = Font(name="Arial", bold=bold, color=fg or "000000",
                     size=size, italic=italic)
    if bg:
        cell.fill = hfill(bg)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    cell.border = thin_border


def apply_block_border(ws, r1, r2, c1, c2):
    """Draw a medium border around the rectangle r1:r2, c1:c2."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            left   = thick if c == c1 else cell.border.left
            right  = thick if c == c2 else cell.border.right
            top    = thick if r == r1 else cell.border.top
            bottom = thick if r == r2 else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def safe_num(val) -> float | None:
    """Cast to float if possible, else None."""
    if val is None:
        return None
    try:
        return round(float(val), 4)
    except (TypeError, ValueError):
        return None


def write_energy_balance_sheet(df_combined: pl.DataFrame,
                               output_path: str=None,
                               units: pl.DataFrame = None,
                               sheet_name: str = "Emissies en energiebalansen",
                               existing_path: str = None,
                               wb: Workbook = None):
    if wb is None:
        if existing_path:
            wb = load_workbook(existing_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

    # Replace sheet if it already exists, otherwise create it
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── column groups ──────────────────────────────────────────────────────
    meta_cols    = META_COLS_ORDER
    emission_cols = EMISSION_COLS_ORDER
    energy_cols = UTILITY_COLS_ORDER
    # Use canonical energy order, but filter to only columns that exist
    # if len(df_combined) > 0:
    #     available_cols = set(df_combined.columns)
    #     energy_cols = [c for c in UTILITY_COLS_ORDER if c in available_cols]
    # else:
    #     energy_cols = UTILITY_COLS_ORDER  # Use full default order
 
    emission_display = {
        "CO2":     "CO₂ emissies scope 1\n(volg NEa richtlijn)",
        "Methane": "Methaan scope 1\nemissies",
        "N2O":     "N₂O scope 1 emissies",
        "F-gases": "F-gassen scope 1\nemissies",
        "CO2 (fossil) CCU/CCS": "CO2 (fossil) CCU/CCS",
        "CO2 (bio) CCU/CCS": "CO2 (bio) CCU/CCS",
    }

    # orange_headers = {"CO2", "Methane", "N2O", "F-gases", "CO2 (fossil) CCU/CCS", "CO2 (bio) CCU/CCS"}
    orange_headers = set(EMISSION_COLS_ORDER)
 

    def energy_display(col):
        return col.replace("_peak", "\n peak").replace("_", " ")

    em_start = 4
    em_end   = em_start + len(emission_cols) - 1
    en_start = em_end + 1
    en_end   = en_start + len(energy_cols) - 1

    # ── ROW 1: section title ───────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    c = ws.cell(1, 1, "2. Emissies en energiebalansen")
    style_cell(c, bold=True, size=14, align="left", fg=WHITE, bg=MID_BLUE)
    for col in range(2, 4):
        ws.cell(1, col).fill = hfill(MID_BLUE)

    # ── ROW 2: group headers ───────────────────────────────────────────────
    for col in range(1, 4):
        ws.cell(2, col).fill = hfill(NAVY)

    ws.merge_cells(start_row=2, start_column=em_start, end_row=2, end_column=em_end)
    c = ws.cell(2, em_start, "Broeikasgas emissies en CC(U)S")
    style_cell(c, bold=True, fg=WHITE, bg=NAVY, size=10)

    ws.merge_cells(start_row=2, start_column=en_start, end_row=2, end_column=en_end)
    c = ws.cell(2, en_start, "Energiebalans")
    style_cell(c, bold=True, fg=WHITE, bg=NAVY, size=10)

    # ── ROW 3: sub-headers ─────────────────────────────────────────────────
    for col in range(1, 4):
        ws.cell(3, col).fill = hfill(NAVY)

    for i, ec in enumerate(emission_cols):
        c = ws.cell(3, em_start + i, emission_display.get(ec, ec))
        bg = ORANGE if ec in orange_headers else HEADER_BLUE
        fg = "000000" if ec in orange_headers else WHITE
        style_cell(c, bold=True, fg=fg, bg=bg, wrap=True, size=9)

    for i, nc in enumerate(energy_cols):
        c = ws.cell(3, en_start + i, energy_display(nc))
        style_cell(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)

    # ── ROW 4: units ──────────────────────────────────────────────────────
    for col in range(1, 4):
        ws.cell(4, col).fill = hfill(NAVY)

    for i, ec in enumerate(emission_cols):
        c = ws.cell(4, em_start + i, get_unit(units, ec))
        bg = ORANGE if ec in orange_headers else HEADER_BLUE
        fg = "000000" if ec in orange_headers else WHITE
        style_cell(c, fg=fg, bg=bg, size=8, italic=True)

    for i, nc in enumerate(energy_cols):
        if 'peak' in nc.lower():
            is_peak = True
            nc = 'Electricity'
        else:
            is_peak = False
        c = ws.cell(4, en_start + i, get_unit(units, nc, is_peak))
        style_cell(c, fg=WHITE, bg=HEADER_BLUE, size=8, italic=True)

    # ── DATA ROWS ──────────────────────────────────────────────────────────
    all_cols = meta_cols + emission_cols + energy_cols
    
    current_row = 5
    year_order = [REFERENCE_YEAR] + SCENARIO_YEARS
    flow_order = FLOW_TYPES
    strategies_to_write = ["Reference"] + [s for s in STRATEGIES_ORDER[1:]]
    
    if len(df_combined) > 0:
        data = (
            df_combined
            .with_columns(pl.col("Year").cast(pl.Utf8))
            .select([c for c in all_cols if c in df_combined.columns])
            .to_dicts()
        )
 
        # group: strategy → year (preserving insertion order)
        rows_by_strategy = {}
        for row in data:
            rows_by_strategy.setdefault(row["Strategy"], []).append(row)
 
        # strategies_to_write = list(rows_by_strategy.keys())
    else:
        # If no data, still write structure with empty values
        data = []
        
        rows_by_strategy = {}
    for strategy in strategies_to_write:
        s_rows = rows_by_strategy.get(strategy, [])

        rows_by_year = {}
        for row in s_rows:
            rows_by_year.setdefault(str(row["Year"]), []).append(row)

        first_year = True

        for year in year_order:
            # Reference strategy only contains the reference year
            if strategy == "Reference" and str(year) != str(REFERENCE_YEAR):
                continue

            # All other strategies skip the reference year
            if strategy != "Reference" and str(year) == str(REFERENCE_YEAR):
                continue

            year_start_row = current_row

            existing_rows = {
                str(r.get("Flow type", "")).lower(): r
                for r in rows_by_year.get(str(year), [])
            }

            year_rows = []

            for flow in FLOW_TYPES:
                row = existing_rows.get(flow.lower())

                if row is None:
                    row = {
                        "Year": year,
                        "Strategy": strategy,
                        "Flow type": flow.capitalize(),
                    }

                year_rows.append(row)

            for r_idx, row in enumerate(year_rows):
                flow = row.get("Flow type", "")
                is_supply = flow.lower() == "supply"
                is_production = flow.lower() == "production"
                row_bg = GREY_ROW if (is_supply or is_production) else WHITE
 
                # col A: strategy — only on the very first row of this strategy
                show_strategy = first_year and r_idx == 0
                c = ws.cell(current_row, 1, strategy if show_strategy else "")
                style_cell(c, bold=show_strategy, fg="000000",
                        bg=LIGHT_BLUE if show_strategy else row_bg,
                        align="left", size=11 if show_strategy else 10)
 
                # col B: year — written on first row, merged later
                c = ws.cell(current_row, 2, year if r_idx == 0 else "")
                style_cell(c, bold=True, fg="000000",
                        bg=LIGHT_BLUE if show_strategy else row_bg,
                        align="left", valign="top")
 
                # col C: flow type
                c = ws.cell(current_row, 3, flow.capitalize())
                style_cell(c, fg="000000", bg=row_bg, align="left")
 
                # emission values
                for i, ec in enumerate(emission_cols):
                    val = row.get(ec)
                    c = ws.cell(current_row, em_start + i,
                                round(val, 4) if val is not None else None)
                    style_cell(c, bg=row_bg, align="right")
                    if val is not None:
                        c.number_format = '#,##0.00'
 
                # energy values
                for i, nc in enumerate(energy_cols):
                    val = row.get(nc)
                    c = ws.cell(current_row, en_start + i,
                                round(val, 4) if val is not None else None)
                    style_cell(c, bg=row_bg, align="right")
                    if val is not None:
                        c.number_format = '#,##0.00'
 
                current_row += 1
 
            year_end_row = current_row - 1
 
            # merge year cell (col B) across all rows of this year block
            if year_end_row > year_start_row:
                ws.merge_cells(
                    start_row=year_start_row, start_column=2,
                    end_row=year_end_row,     end_column=2
                )
                # re-apply style to merged cell so text + bg survive the merge
                c = ws.cell(year_start_row, 2)
                style_cell(c, bold=True, fg="000000",
                        bg=LIGHT_BLUE if first_year else WHITE,
                        align="left", valign="top")
 
            # thick border around the whole year block (col B → last energy col)
            apply_block_border(ws, year_start_row, year_end_row, 2, en_end)
 
            first_year = False
 
        # blank separator row between strategy groups
        for col in range(1, en_end + 1):
            ws.cell(current_row, col).fill = hfill(WHITE)
        current_row += 1
 
    # ── column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22   # strategy
    ws.column_dimensions["B"].width = 10   # year
    ws.column_dimensions["C"].width = 18   # flow type
    for i in range(len(emission_cols)):
        ws.column_dimensions[get_column_letter(em_start + i)].width = 16
    for i in range(len(energy_cols)):
        ws.column_dimensions[get_column_letter(en_start + i)].width = 14
 
    # row heights
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[4].height = 16
 
    ws.freeze_panes = "D5"
    
    if output_path:
        wb.save(output_path)
        print(f"Saved to {output_path}")
    return wb

def write_plant_details_sheet(
    df: pl.DataFrame,
    output_path: str = None,
    sheet_name: str = "Plant details",
    existing_path: str = None,
    wb: Workbook = None
):
    """
    Writes a two-column (Field, Value) plant details sheet.
    Field column: bold, light blue background.
    Value column: bold, light gray background.
    Thick borders between all cells.

    Can work in two modes:
    1. File mode: provide output_path (existing behavior)
    2. Workbook mode: provide wb (for Streamlit)
    """

    if wb is None:
        # load or create workbook
        if existing_path:
            wb = load_workbook(existing_path)
        else:
            wb = Workbook()
            # Remove the default empty sheet Workbook() always creates
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
 
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
 
    FIELD_BG = "BDD7EE"  # light blue
    VALUE_BG  = "D9D9D9"  # light gray
    border = Border(
        left=thick, right=thick, top=thick, bottom=thick
    )
 
    def write_cell(row, col, value, bg):
        c = ws.cell(row, col, value)
        c.font      = Font(name="Arial", bold=True, size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.border    = border
 
    rows = df.to_dicts()
    field_col, value_col = df.columns[0], df.columns[1]
 
    max_field_len = 0
    max_value_len = 0
 
    for r_idx, row in enumerate(rows, start=1):
        field_val = row[field_col]
        value_val = row[value_col]
        write_cell(r_idx, 1, field_val, FIELD_BG)
        write_cell(r_idx, 2, value_val, VALUE_BG)
        ws.row_dimensions[r_idx].height = 18
        max_field_len = max(max_field_len, len(str(field_val or "")))
        max_value_len = max(max_value_len, len(str(value_val or "")))
 
    # Autosize: character length * ~1.2 scaling + small padding
    ws.column_dimensions["A"].width = min(max_field_len * 1.2 + 4, 60)
    ws.column_dimensions["B"].width = min(max_value_len * 1.2 + 4, 80)
 
    if output_path:
        wb.save(output_path)
    
    return wb


def write_projects_sheet(
    df: pl.DataFrame,
    output_path: str = None,
    units: pl.DataFrame = None,
    sheet_name: str = "Projecten",
    existing_path: str = None,
    wb: Workbook = None
):
    """
    Writes the projects sheet. Each project spans N flow-type rows (Delta demand,
    Delta captive use, Delta production, Delta supply). Project name + detail
    columns are merged vertically across those rows.
    """

    if wb is None:
        if existing_path:
            wb = load_workbook(existing_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
 
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
 
    # ── column definitions ─────────────────────────────────────────────────
    # Project name is its own frozen column; remaining details start at col 4
    PROJ_NAME_COL_NAME = "Project name"
    # DETAIL_COLS = [
    #     "Project Type", "Description", "Project phase",
    #     "Prob. of success", "Year of operation", "Planned execute year",
    #     "Planned define year", "Part of Preferred Strategy",
    #     "Associated Strategies", "Associated Scenarios", "EAN",
    # ]

    DETAIL_COLS = PROJECT_DETAILS_ORDER + ['EAN']
    if 'CO2' in DETAIL_COLS:
        DETAIL_COLS.remove('CO2')

    FLOW_COL    = "Type"
    FLOW_TYPES  = ["Delta demand", "Delta captive use", "Delta production", "Delta supply"]
 
    EMISSION_COLS = ["CO2", "CO2 (fossil) CCU/CCS", "CO2 (bio) CCU/CCS"]
    VALUE_COLS = [c for c in df.columns
                  if c not in DETAIL_COLS + [FLOW_COL] + EMISSION_COLS + [PROJ_NAME_COL_NAME]]
 
    # ── colours ────────────────────────────────────────────────────────────
    NAVY        = "1F3864"
    WHITE       = "FFFFFF"
    LIGHT_BLUE  = "BDD7EE"
    GREY_ROW    = "F2F2F2"
    HEADER_BLUE = "4472C4"
    MID_BLUE    = "2E75B6"
 
    thick_s = Side(style="medium", color="000000")
    thin_s  = Side(style="thin",   color="CCCCCC")
    thin_border  = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
 
    def hfill(c): return PatternFill("solid", fgColor=c)
 
    def sc(cell, bold=False, fg="000000", bg=None, align="center",
           valign="center", size=10, wrap=False, italic=False):
        cell.font      = Font(name="Arial", bold=bold, color=fg, size=size, italic=italic)
        cell.fill      = hfill(bg) if bg else PatternFill()
        cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
        cell.border    = thin_border
 
    def block_border(r1, r2, c1, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                cell.border = Border(
                    left   = thick_s if c == c1 else cell.border.left,
                    right  = thick_s if c == c2 else cell.border.right,
                    top    = thick_s if r == r1 else cell.border.top,
                    bottom = thick_s if r == r2 else cell.border.bottom,
                )
 
    # ── column index map ──────────────────────────────────────────────────
    # Layout: proj# | project name | flow type | detail cols | emission cols | value cols
    proj_num_col  = 1
    proj_name_col = 2
    flow_col_idx  = 3
    detail_start  = 4
    detail_end    = detail_start + len(DETAIL_COLS) - 1
    em_start      = detail_end + 1
    em_end        = em_start + len(EMISSION_COLS) - 1
    val_start     = em_end + 1
    val_end       = val_start + len(VALUE_COLS) - 1
    total_cols    = val_end
 
    # ── ROW 1: title ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, "3. Projecten")
    sc(c, bold=True, size=14, fg=WHITE, bg=MID_BLUE, align="left")
    for col in range(2, total_cols + 1):
        ws.cell(1, col).fill = hfill(MID_BLUE)
 
    # ── ROW 2: group headers ──────────────────────────────────────────────
    for col in range(1, total_cols + 1):
        ws.cell(2, col).fill = hfill(NAVY)
 
    # frozen cols group: proj# + project name + flow type
    ws.merge_cells(start_row=2, start_column=proj_num_col,
                   end_row=2, end_column=flow_col_idx)
    c = ws.cell(2, proj_num_col, "Project details")
    sc(c, bold=True, fg=WHITE, bg=NAVY)
 
    # remaining detail cols
    ws.merge_cells(start_row=2, start_column=detail_start,
                   end_row=2, end_column=detail_end)
    c = ws.cell(2, detail_start, "")
    sc(c, bg=NAVY)
 
    ws.merge_cells(start_row=2, start_column=em_start,
                   end_row=2, end_column=em_end)
    c = ws.cell(2, em_start, "Broeikasgas emissies en CC(U)S")
    sc(c, bold=True, fg=WHITE, bg=HEADER_BLUE)
 
    ws.merge_cells(start_row=2, start_column=val_start,
                   end_row=2, end_column=val_end)
    c = ws.cell(2, val_start, "Energiebalans")
    sc(c, bold=True, fg=WHITE, bg=NAVY)
 
    # ── ROW 3: sub-headers ────────────────────────────────────────────────
    # header_display = {
    #     "Project name": "Project name", "Project Type": "Project type",
    #     "Description": "Description", "Project phase": "Phase",
    #     "Prob. of success": "Probability of success",
    #     "Year of operation": "Year of operation",
    #     "Planned execute year": "Planned execute year",
    #     "Planned define year": "Planned define year",
    #     "Part of Preferred Strategy": "Part of preferred strategy",
    #     "Associated Strategies": "Associated Strategies",
    #     "Associated Scenarios": "Associated Scenarios",
    #     "EAN": "EAN",
    # }
    c = ws.cell(3, proj_num_col, "#")
    sc(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
    c = ws.cell(3, proj_name_col, "Project name")
    sc(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
    c = ws.cell(3, flow_col_idx, "Flow type")
    sc(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
    for i, col in enumerate(DETAIL_COLS):
        c = ws.cell(3, detail_start + i, col)
        sc(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
    for i, col in enumerate(EMISSION_COLS):
        c = ws.cell(3, em_start + i, col)
        sc(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
    for i, col in enumerate(VALUE_COLS):
        c = ws.cell(3, val_start + i, col)
        sc(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
    # ── ROW 4: units ──────────────────────────────────────────────────────
    for col in range(1, total_cols + 1):
        ws.cell(4, col).fill = hfill(NAVY)

    for i, ec in enumerate(EMISSION_COLS):
        c = ws.cell(4, em_start + i, get_unit(units, ec))
        sc(c, fg=WHITE, bg=HEADER_BLUE, size=8, italic=True)

    for i, vc in enumerate(VALUE_COLS):
        is_peak = 'peak' in vc.lower()
        lookup  = 'Electricity' if is_peak else vc
        c = ws.cell(4, val_start + i, get_unit(units, lookup, is_peak))  
        sc(c, fg=WHITE, bg=HEADER_BLUE, size=8, italic=True)
 
    # ── DATA ROWS ─────────────────────────────────────────────────────────
    current_row = 5
    data = df.to_dicts()
 
    # group by project name, preserving order
    rows_by_project = {}
    for row in data:
        rows_by_project.setdefault(row["Project name"], []).append(row)
 
    proj_counter = 0
    for project, p_rows in rows_by_project.items():
        proj_counter += 1
        proj_start = current_row
 
        # sort flow types in display order
        flow_order = {f: i for i, f in enumerate(FLOW_TYPES)}
        p_rows_sorted = sorted(p_rows, key=lambda r: flow_order.get(r.get(FLOW_COL, ""), 99))
 
        for r_idx, row in enumerate(p_rows_sorted):
            flow  = row.get(FLOW_COL, "")
            is_even = r_idx % 2 == 1
            row_bg = GREY_ROW if is_even else WHITE
 
            is_first = r_idx == 0
 
            # col 1: project number — merged vertically
            c = ws.cell(current_row, proj_num_col, f"Project {proj_counter}" if is_first else "")
            sc(c, bold=True, fg="000000", bg=LIGHT_BLUE if is_first else row_bg,
               align="center", valign="top" if is_first else "center")
 
            # col 2: project name — merged vertically
            c = ws.cell(current_row, proj_name_col, project if is_first else "")
            sc(c, bold=True, fg="000000", bg=LIGHT_BLUE if is_first else row_bg,
               align="left", wrap=True, valign="top" if is_first else "center")
 
            # col 3: flow type — one per row, NOT merged
            c = ws.cell(current_row, flow_col_idx, flow)
            sc(c, fg="000000", bg=row_bg, align="left")
 
            # remaining detail columns — merged vertically
            for i, col in enumerate(DETAIL_COLS):
                val = row.get(col) if is_first else ""
                c = ws.cell(current_row, detail_start + i, val)
                sc(c, fg="000000", bg=LIGHT_BLUE if is_first else row_bg,
                   align="left", wrap=True, valign="top" if is_first else "center")
 
            # emission values
            for i, col in enumerate(EMISSION_COLS):
                show = flow == "Delta production" or col != "CO2"
                val = safe_num(row.get(col)) if show else None
                # val = safe_num(row.get(col))
                c = ws.cell(current_row, em_start + i, val)
                sc(c, bg=row_bg, align="right")
                if val is not None:
                    c.number_format = '#,##0.00'
 
            # energy/value cols
            for i, col in enumerate(VALUE_COLS):
                val = safe_num(row.get(col))
                c = ws.cell(current_row, val_start + i, val)
                sc(c, bg=row_bg, align="right")
                if val is not None:
                    c.number_format = '#,##0.00'
 
            current_row += 1
 
        proj_end = current_row - 1
 
        # merge proj#, project name, and detail cols vertically (NOT flow type)
        if proj_end > proj_start:
            for merge_col, kw in [
                (proj_num_col, dict(bold=True, align="center")),
                (proj_name_col, dict(bold=True, align="left", wrap=True)),
            ]:
                ws.merge_cells(start_row=proj_start, start_column=merge_col,
                               end_row=proj_end,     end_column=merge_col)
                c = ws.cell(proj_start, merge_col)
                sc(c, fg="000000", bg=LIGHT_BLUE, valign="top", **kw)
 
            for i in range(len(DETAIL_COLS)):
                col_idx = detail_start + i
                ws.merge_cells(start_row=proj_start, start_column=col_idx,
                               end_row=proj_end,     end_column=col_idx)
                c = ws.cell(proj_start, col_idx)
                sc(c, fg="000000", bg=LIGHT_BLUE, align="left",
                   wrap=True, valign="top")
 
        # thick border around entire project block
        block_border(proj_start, proj_end, proj_num_col, total_cols)
 
        # blank separator row
        for col in range(1, total_cols + 1):
            ws.cell(current_row, col).fill = hfill(WHITE)
        current_row += 1
 
    # ── column widths ─────────────────────────────────────────────────────
    col_widths = {
        1: 12,  # project number
        2: 24,  # project name
        3: 18,  # flow type
        4: 14, 5: 20, 6: 12, 7: 14, 8: 12, 9: 14,
        10: 14, 11: 20, 12: 20, 13: 12, 14: 12,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for i in range(len(EMISSION_COLS)):
        ws.column_dimensions[get_column_letter(em_start + i)].width = 16
    for i in range(len(VALUE_COLS)):
        ws.column_dimensions[get_column_letter(val_start + i)].width = 14
 
    ws.row_dimensions[3].height = 36
    # Freeze cols 1-3: proj#, project name, flow type
    ws.freeze_panes = f"{get_column_letter(flow_col_idx + 1)}5"  # = D5
 
    if output_path:
        wb.save(output_path)
    
    return wb


# for storage, flexibility and production

def _write_simple_sheet(ws, df: pl.DataFrame):
    """Shared formatting: bold light-blue headers, borders on all cells, autosized columns."""
    LIGHT_BLUE = "BDD7EE"
    thin_s = Side(style="thin", color="000000")
    border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
 
    def cell_style(cell, bold=False, bg=None):
        cell.font = Font(name="Arial", bold=bold, size=10)
        cell.fill = PatternFill("solid", fgColor=bg) if bg else PatternFill()
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
 
    # Write headers even if df is empty
    for c_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(1, c_idx, col_name)
        cell_style(c, bold=True, bg=LIGHT_BLUE)
        
        if len(df) > 0:
            col_data = df[col_name].cast(pl.Utf8, strict=False).drop_nulls().to_list()
            max_len = max([len(col_name)] + [len(str(v)) for v in col_data], default=10)
        else:
            max_len = len(col_name)
            
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len * 1.1 + 2, 60)
 
    # Write data rows only if not empty
    if len(df) > 0:
        for r_idx, row in enumerate(df.to_dicts(), start=2):
            for c_idx, col_name in enumerate(df.columns, start=1):
                c = ws.cell(r_idx, c_idx, row.get(col_name))
                cell_style(c)
 
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18
 
 
def _simple_sheet_writer(sheet_name: str):
    """Factory that returns a write function for a simple formatted sheet."""
    def write_fn(
        df: pl.DataFrame,
        output_path: str = None,
        sheet_name: str = sheet_name,
        existing_path: str = None,
        wb: Workbook = None
    ):
        if wb is None:
            if existing_path:
                wb = load_workbook(existing_path)
            else:
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        _write_simple_sheet(ws, df)

        if output_path:
            wb.save(output_path)
            print(f"Saved '{sheet_name}' to {output_path}")
    
        return wb
        
    write_fn.__name__ = f"write_{sheet_name.lower()}_sheet"
    return write_fn
 
 
write_production_sheet   = _simple_sheet_writer("Production")
write_storage_sheet      = _simple_sheet_writer("Storage")
write_flexibility_sheet  = _simple_sheet_writer("Flexibility")
    

def write_scenario_sheets(
    df_combined: pl.DataFrame,
    output_path: str=None,
    existing_path: str = None,
    n_scenarios: int = 5,
    ref_year: int = 2024,
    units: pl.DataFrame = None,
    wb: Workbook = None,
    scenario_names: dict[int, str]= SCENARIO_NAMES_DICT,
    scenario_years: list[str] = SCENARIO_YEARS
):
    """
    Writes n_scenarios template sheets ("Scenario 1" … "Scenario n").
    Same header/structure as write_sheet but:
      - No reference year rows
      - Single strategy placeholder "strategy" (merged across all years)
      - All emission and energy value cells empty
    """
    if wb is None:
        if existing_path:
            wb = load_workbook(existing_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
 
    # ── derive column lists from the real df (same as write_sheet) ─────────
    meta_cols     = META_COLS_ORDER
    emission_cols = EMISSION_COLS_ORDER
    energy_cols = UTILITY_COLS_ORDER

    # if len(df_combined) > 0:
    #     energy_cols   = [c for c in df_combined.columns
    #                      if c not in meta_cols + emission_cols]
    # else:
    #     energy_cols = ["Electricity", "Natural Gas", "Hydrogen"]
        
    emission_display = {
        "CO2":     "CO₂ emissies scope 1\n(volg NEa richtlijn)",
        "Methane": "Methaan scope 1\nemissies",
        "N2O":     "N₂O scope 1 emissies",
        "F-gases": "F-gassen scope 1\nemissies",
        "CO2 (fossil) CCU/CCS": "CO2 (fossil) CCU/CCS",
        "CO2 (bio) CCU/CCS":    "CO2 (bio) CCU/CCS",
    }
 
    def energy_display(col):
        return col.replace("_peak", "\n(peak)").replace("_", " ")
 
    em_start = 4
    em_end   = em_start + len(emission_cols) - 1
    en_start = em_end + 1
    en_end   = en_start + len(energy_cols) - 1
 
    year_cols  = scenario_years
    flow_types = FLOW_TYPES
 
    def build_sheet(ws, scenario_num: int, ref_data: list):
        # ── ROW 1: title ──────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        c = ws.cell(1, 1, f"Scenario {scenario_names[scenario_num]}")
        style_cell(c, bold=True, size=14, align="left", fg=WHITE, bg=MID_BLUE)
        for col in range(2, 4):
            ws.cell(1, col).fill = hfill(MID_BLUE)
 
        # ── ROW 2: group headers ──────────────────────────────────────────
        for col in range(1, 4):
            ws.cell(2, col).fill = hfill(NAVY)
 
        ws.merge_cells(start_row=2, start_column=em_start, end_row=2, end_column=em_end)
        c = ws.cell(2, em_start, "Broeikasgas emissies en CC(U)S")
        style_cell(c, bold=True, fg=WHITE, bg=NAVY, size=10)
 
        ws.merge_cells(start_row=2, start_column=en_start, end_row=2, end_column=en_end)
        c = ws.cell(2, en_start, "Energiebalans")
        style_cell(c, bold=True, fg=WHITE, bg=NAVY, size=10)
 
        # ── ROW 3: sub-headers ────────────────────────────────────────────
        for col in range(1, 4):
            ws.cell(3, col).fill = hfill(NAVY)
 
        for i, ec in enumerate(emission_cols):
            c = ws.cell(3, em_start + i, emission_display.get(ec, ec))
            style_cell(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
        for i, nc in enumerate(energy_cols):
            c = ws.cell(3, en_start + i, energy_display(nc))
            style_cell(c, bold=True, fg=WHITE, bg=HEADER_BLUE, wrap=True, size=9)
 
        # ── ROW 4: units ──────────────────────────────────────────────────
        for col in range(1, 4):
            ws.cell(4, col).fill = hfill(NAVY)
 
        for i, ec in enumerate(emission_cols):
            unit = get_unit(units, ec) if units is not None else None
            c = ws.cell(4, em_start + i, unit)
            style_cell(c, fg=WHITE, bg=HEADER_BLUE, size=8, italic=True)

        for i, nc in enumerate(energy_cols):
            is_peak = 'peak' in nc.lower()
            lookup = 'Electricity' if is_peak else nc
            unit = get_unit(units, lookup, is_peak) if units is not None else None
            c = ws.cell(4, en_start + i, unit)
            style_cell(c, fg=WHITE, bg=HEADER_BLUE, size=8, italic=True)

        # ── DATA ROWS ─────────────────────────────────────────────────────
        current_row = 5
 
        def write_year_block(year, strategy_label, rows_dict, is_template):
            """Write one year block. If is_template, values are empty."""
            nonlocal current_row
            year_start_row = current_row
            year_flows = flow_types  # always 4 flow types
 
            for r_idx, flow in enumerate(year_flows):
                is_even = r_idx % 2 == 1
                row_bg  = GREY_ROW if is_even else WHITE
 
                c = ws.cell(current_row, 1, strategy_label if r_idx == 0 else "")
                style_cell(c, bold=r_idx == 0, fg="000000", bg=row_bg, align="left")
 
                c = ws.cell(current_row, 2, year if r_idx == 0 else "")
                style_cell(c, bold=True, fg="000000", bg=row_bg,
                           align="left", valign="top")
 
                c = ws.cell(current_row, 3, flow.capitalize())
                style_cell(c, fg="000000", bg=row_bg, align="left")
 
                # values: real data if reference, empty if template
                ref_row = rows_dict.get(flow.lower()) if rows_dict else None
                for i, ec in enumerate(emission_cols):
                    val = (ref_row.get(ec) if ref_row and not is_template else None)
                    c = ws.cell(current_row, em_start + i,
                                round(val, 4) if val is not None else None)
                    style_cell(c, bg=row_bg, align="right")
                    if val is not None:
                        c.number_format = "#,##0.00"
 
                for i, nc in enumerate(energy_cols):
                    val = (ref_row.get(nc) if ref_row and not is_template else None)
                    c = ws.cell(current_row, en_start + i,
                                round(val, 4) if val is not None else None)
                    style_cell(c, bg=row_bg, align="right")
                    if val is not None:
                        c.number_format = "#,##0.00"
 
                current_row += 1
 
            year_end_row = current_row - 1
 
            if year_end_row > year_start_row:
                ws.merge_cells(start_row=year_start_row, start_column=2,
                               end_row=year_end_row,     end_column=2)
                c = ws.cell(year_start_row, 2)
                style_cell(c, bold=True, fg="000000", bg=WHITE,
                           align="left", valign="top")
 
            apply_block_border(ws, year_start_row, year_end_row, 2, en_end)
            return year_start_row, year_end_row
 
        # ── reference year block (real data) ──────────────────────────────
        ref_rows = {
            row["Flow type"].lower(): row
            for row in ref_data
            if str(row["Year"]) == str(ref_year)
        }
        ref_start, _ = write_year_block(str(ref_year), "Reference", ref_rows, is_template=False)
        # blank separator
        for col in range(1, en_end + 1):
            ws.cell(current_row, col).fill = hfill(WHITE)
        current_row += 1
 
        # ── template years (empty values) ─────────────────────────────────
        strategy_start = current_row
        for year in year_cols:
            label = "strategy" if year == year_cols[0] else ""
            write_year_block(year, label, None, is_template=True)
 
        # merge strategy col across template rows
        strategy_end = current_row - 1
        if strategy_end > strategy_start:
            ws.merge_cells(start_row=strategy_start, start_column=1,
                           end_row=strategy_end,     end_column=1)
        c = ws.cell(strategy_start, 1)
        style_cell(c, bold=True, fg="000000", bg=LIGHT_BLUE,
                   align="left", valign="top")
 
        # also style reference strategy cell
        c = ws.cell(ref_start, 1)
        style_cell(c, bold=True, fg="000000", bg=LIGHT_BLUE,
                   align="left", valign="top")
 
        # ── column widths ─────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 18
        for i in range(len(emission_cols)):
            ws.column_dimensions[get_column_letter(em_start + i)].width = 16
        for i in range(len(energy_cols)):
            ws.column_dimensions[get_column_letter(en_start + i)].width = 14
 
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 40
        ws.row_dimensions[4].height = 16
        ws.freeze_panes = "D5"
 
    # ── extract reference rows from df_combined ───────────────────────────
    ref_data = (
        df_combined
        .with_columns(pl.col("Year").cast(pl.Utf8))
        .filter(pl.col("Strategy") == "Reference")
        .to_dicts()
    )
 
    # ── write all scenario sheets ──────────────────────────────────────────
    for n in range(1, n_scenarios + 1):
        sheet_name = f"Scenario {scenario_names[n]}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        build_sheet(ws, n, ref_data)
 
    if output_path:
        wb.save(output_path)
    
    return wb
