import polars as pl
import pandas as pd
import io
from openpyxl import Workbook, load_workbook
import tempfile
from pathlib import Path
from.constants import REFERENCE_YEAR
from .utils import (parse_All_EANs, 
                    get_plant_name, 
                    get_all_units, 
                    get_production_sheet, 
                    get_energy_balance_emissions_sheet, 
                    get_project_sheet, 
                    get_storage_sheet, 
                    get_flexibility_sheet, 
                    create_company_details_dfs
)

from .format_emissies_sheet import (write_energy_balance_sheet, 
                                    write_flexibility_sheet, 
                                    write_plant_details_sheet, 
                                    write_production_sheet, 
                                    write_projects_sheet, 
                                    write_scenario_sheets, 
                                    write_storage_sheet)

from CONNECT_CTM.read_DSH_files import read_all_scenario_sheets

def read_csv_streamlit(streamlit_file) -> pl.DataFrame:
    """Read a CSV via pandas, falling back to Polars for malformed files."""
    try:
        return pl.from_pandas(pd.read_csv(streamlit_file))
    except pd.errors.ParserError:
        print(f"  [warn] Falling back to Polars for {streamlit_file}")
        return pl.read_csv(streamlit_file, truncate_ragged_lines=True)

def read_xlsx_streamlit(streamlit_file) -> pl.DataFrame:
    """Read a xlsx via polars, falling back to pandas for malformed files."""
    try:
        return pl.read_excel(streamlit_file, truncate_ragged_lines=True)
    except:
        print(f"  [warn] Falling back to Polars for {streamlit_file}")
        return pl.from_pandas(pd.read_excel(streamlit_file))


def load_data_streamlit(uploaded_files:dict) -> dict[str, pl.DataFrame]:
    """Load all source CSVs and return as a named dict."""
    print("Loading data...")
    data = {name: read_csv_streamlit(filename) for name, filename in uploaded_files.items()}
    # Parse EANs once here so it's not repeated per plant
    data["plants_parsed"] = parse_All_EANs(data["plants"])
    print(f"  Loaded {len(uploaded_files)} files.")
    return data


def process_plant_streamlit(
    plant_id: str, 
    data: dict[str, pl.DataFrame],
    logs: list = None,
) -> tuple[io.BytesIO, list]:
    """
    Process a single plant and generate Excel in memory.
    
    Args:
        plant_id: Plant ID
        data: Dict with all DataFrames (same as before)
        logs: List to append logs to
    
    Returns:
        (BytesIO object, logs list)
    """
    
    if logs is None:
        logs = []
    
    plant_name = get_plant_name(data["plants"], plant_id)
    logs.append(f"Processing: {plant_name}")
    
    try:
        # Get units
        units = get_all_units(
            ref_emission_df=data["emission_reference"],
            ref_utility_df=data["demand_reference"],
        )
        
        # ── Data preparation ───────────────────────────────────────────
        energy_balance = get_energy_balance_emissions_sheet(
            data["emission_forecast"],
            data["emission_reference"],
            data["demand_forecast"],
            data["demand_reference"],
            plant_id,
            REFERENCE_YEAR,
        )
        
        details = create_company_details_dfs(
            data["plants"],
            plant_id,
            parsed_df=data["plants_parsed"],
        )
        
        projects = get_project_sheet(
            data["project_emissions"],
            data["project_utilities"],
            plant_id,
        )
        
        production = get_production_sheet(data["production"], plant_id)
        storage = get_storage_sheet(data["storage"], plant_id)
        flexibility = get_flexibility_sheet(data["flexibility"], plant_id)
        
        # ── Create workbook in memory ──────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # ── Write sheets to workbook ──────────────────────────────────
        wb = write_plant_details_sheet(details, wb=wb)
        wb = write_energy_balance_sheet(energy_balance, wb=wb, units=units)
        wb = write_projects_sheet(projects, wb=wb, units=units)
        wb = write_production_sheet(production, wb=wb)
        wb = write_storage_sheet(storage, wb=wb)
        wb = write_flexibility_sheet(flexibility, wb=wb)
        wb = write_scenario_sheets(energy_balance, wb=wb, units=units)
        
        # ── Save to BytesIO ────────────────────────────────────────────
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        logs.append(f"  Generated: {plant_name}.xlsx")
        
        return excel_bytes, logs
    
    except Exception as e:
        logs.append(f"  ERROR: {plant_name}: {e}")
        return None, logs


def extract_scenario_years(streamlit_file) -> tuple[list[str], list[str]]:
    """
    Extract scenario names and years from uploaded Excel file.
    
    Returns:
        (scenario_names, scenario_years)
    """
    # Save to temp
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(streamlit_file.read())
        tmp.flush()
        tmp_path = tmp.name
    
    try:
        wb = load_workbook(tmp_path)
        
        # Get scenario names from sheet names
        scenarios = [
            sheet.replace("Scenario ", "").strip() 
            for sheet in wb.sheetnames 
            if sheet.startswith("Scenario")
        ]
        
        # Get years from first scenario sheet
        if scenarios:
            ws = wb[f"Scenario {scenarios[0]}"]
            
            # Find header row (look for "Year" and "Flow type")
            header_row = None
            for row_idx in range(1, min(10, ws.max_row + 1)):
                year_val = ws.cell(row_idx, 2).value
                flow_val = ws.cell(row_idx, 3).value
                
                if year_val == "Year" and flow_val == "Flow type":
                    header_row = row_idx
                    break
            
            if header_row is None:
                header_row = 3
            
            # Extract years from column B (year_col = 2)
            years = set()
            for row_idx in range(header_row + 1, ws.max_row + 1):
                year_cell = ws.cell(row_idx, 2).value
                if year_cell is not None:
                    years.add(str(year_cell))
            
            years = sorted(years)
        else:
            years = []
        
        return sorted(scenarios), years
    
    finally:
        Path(tmp_path).unlink()
