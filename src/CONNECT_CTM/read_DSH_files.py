import polars as pl
from openpyxl import load_workbook
from pathlib import Path

def normalize_column_name(name: str) -> str:
    """
    Normalize column names to match variations in Excel files.
    Handles Dutch column names by mapping to English equivalents.
    
    E.g., "Electricity\n(peak)" -> "electricity_peak"
         "CO₂ emissies scope 1\n(volg NEa richtlijn)" -> "co2"
    """
    if name is None:
        return None
    
    name_str = str(name).strip()
    
    # Dutch -> English mapping (do this FIRST)
    dutch_map = {
        "CO₂ emissies scope 1\n(volg NEa richtlijn)": "CO2",
        "CO₂ emissies scope 1\n(volg NEa richtlijn) ": "CO2",  # With trailing space
        "Methaan scope 1\nemissies": "Methane",
        "N₂O scope 1 emissies": "N2O",
        "F-gassen scope 1\nemissies": "F-gases",
    }
    
    # Check if exact match (with flexibility on whitespace)
    for dutch, english in dutch_map.items():
        if dutch.strip() in name_str or name_str in dutch.strip():
            name_str = english
            break
    
    # Then normalize as before
    return (
        name_str
        .lower()
        .replace("\n(peak)", "_peak")
        .replace(" (", "_")
        .replace(")", "")
        .replace(" ", "_")
        .replace("-", "_")
        .replace("₂", "2")  # ← Handle subscript 2
        .replace("₃", "3")  # ← Handle subscript 3
    )
 
def find_column_by_name(
    worksheet,
    header_row: int,
    target_col: str,
) -> int:
    """
    Find column index by normalized name matching.
    Handles variations like "Electricity (peak)" vs "Electricity_peak"
    
    Returns:
        column index (1-based) or None if not found
    """
    target_normalized = normalize_column_name(target_col)
    
    for col_idx in range(1, worksheet.max_column + 1):
        header_val = worksheet.cell(header_row, col_idx).value
        if header_val is None:
            continue
        
        header_normalized = normalize_column_name(header_val)
        if header_normalized == target_normalized:
            return col_idx
    
    return None

def read_scenario_sheet(
    workbook_path: str,
    sheet_name: str,
    emission_cols: list[str],
    energy_cols: list[str],
    reference_year: int,
) -> pl.DataFrame:
    """
    Read a Scenario X sheet created by write_scenario_sheets().
 
    Returns a normalized dataframe with:
        Scenario
        Year
        Flow type
        <emission cols>
        <energy cols>
 
    - Skips the reference year
    - Handles merged year cells
    - Keeps rows even if all values are blank
    - Ignores separator rows
    - Dynamically finds column positions (handles variable column ordering)
    - Matches column names with normalization (handles variations like "Electricity (peak)")
    """
 
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
 
    scenario = sheet_name.replace("Scenario ", "")
 
    records = []
    current_year = None
 
    # ── Find header row ────────────────────────────────────────────────
    header_row = None
    
    # Search for header row by looking for "Year" and "Flow type"
    for row_idx in range(1, min(10, ws.max_row + 1)):
        year_val = ws.cell(row_idx, 2).value
        flow_val = ws.cell(row_idx, 3).value
        
        if year_val == "Year" and flow_val == "Flow type":
            header_row = row_idx
            break
    
    if header_row is None:
        # Fallback: assume header is at row 4
        header_row = 3
 
    # ── Find column indices for each requested column ───────────────────
    year_col = 2  # Usually column B
    flow_type_col = 3  # Usually column C
    
    emission_indices = {}
    for col in emission_cols:
        col_idx = find_column_by_name(ws, header_row, col)
        if col_idx:
            emission_indices[col] = col_idx
    
    energy_indices = {}
    for col in energy_cols:
        col_idx = find_column_by_name(ws, header_row, col)
        if col_idx:
            energy_indices[col] = col_idx
    
    # Log if columns are missing or reordered
    # missing_emissions = [c for c in emission_cols if c not in emission_indices]
    missing_energy = [c for c in energy_cols if c not in energy_indices]
    if missing_energy:
        available_cols = []
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(header_row, col_idx).value
            if header_val and str(header_val).strip() not in ["Year", "Flow type"]:
                available_cols.append(str(header_val))
        
        print(f"[WARN] {sheet_name}: missing or unmatched columns")
        # if missing_emissions:
        #     print(f"  Missing emissions: {missing_emissions}")
        if missing_energy:
            print(f"  Missing energy: {missing_energy}")
        if available_cols:
            print(f"  Available columns in Excel: {available_cols}")
 
    # ── Read data rows ─────────────────────────────────────────────────
    data_start = header_row + 1
    
    for row_idx in range(data_start, ws.max_row + 1):
 
        # Year column (merged vertically)
        year_cell = ws.cell(row_idx, year_col).value
        if year_cell is not None:
            current_year = str(year_cell)
 
        # Flow type column
        flow_type = ws.cell(row_idx, flow_type_col).value
 
        # Skip separator rows
        if flow_type is None:
            continue
 
        # Skip reference year
        if current_year == str(reference_year):
            continue
 
        record = {
            "Scenario": scenario,
            "Year": current_year,
            "Flow type": str(flow_type).lower(),
        }
 
        # Emissions (using found column positions)
        for col, col_idx in emission_indices.items():
            record[col] = ws.cell(row_idx, col_idx).value
 
        # Energy (using found column positions)
        for col, col_idx in energy_indices.items():
            record[col] = ws.cell(row_idx, col_idx).value
 
        # Keep row even if all values are empty
        records.append(record)
 
    if not records:
        return pl.DataFrame(
            schema={
                "Scenario": pl.Utf8,
                "Year": pl.Utf8,
                "Flow type": pl.Utf8,
                **{c: pl.Float64 for c in emission_cols},
                **{c: pl.Float64 for c in energy_cols},
            }
        )
 
    return pl.DataFrame(records)
 
 
def aggregate_scenarios_flow_types(all_df: pl.DataFrame) -> pl.DataFrame:
 
    value_cols = [c for c in all_df.columns if c not in ('Scenario', 'Year', 'Flow type')]
 
    production = all_df.filter(pl.col('Flow type') == 'production')
 
    demand = (
        all_df.filter(pl.col('Flow type').is_in(['demand', 'captive use']))
        .group_by(['Scenario', 'Year'])
        .agg([pl.col(c).sum() for c in value_cols])
        .with_columns(pl.lit('demand').alias('Flow type'))
        .select(all_df.columns)  # restore original column order
    )
 
    result = pl.concat([production, demand])
    return result
 
 
def read_all_scenario_sheets(
    workbook_path: str,
    emission_cols: list[str],
    energy_cols: list[str],
    reference_year: int,
    aggregate_flow_types:bool = True
) -> pl.DataFrame:
    """
    Read all Scenario X sheets and return a single dataframe.
    """
    wb = load_workbook(workbook_path, read_only=True)   
 
    scenario_sheets = [
        sheet
        for sheet in wb.sheetnames
        if sheet.startswith("Scenario ")
    ]
 
    dfs = [
        read_scenario_sheet(
            workbook_path=workbook_path,
            sheet_name=sheet,
            emission_cols=emission_cols,
            energy_cols=energy_cols,
            reference_year=reference_year,
        )
        for sheet in scenario_sheets
    ]

    if not dfs:
        return pl.DataFrame()
    
    if aggregate_flow_types:
        all_df = pl.concat(dfs, how="vertical_relaxed")
        return aggregate_scenarios_flow_types(all_df)
 
    return pl.concat(dfs, how="vertical_relaxed")
 
 
def save_scenario_values(
    workbook_path: str,
    plant_name: str,
    output_dir: str,
    emission_cols: list[str],
    energy_cols: list[str],
    reference_year: int,
) -> str:
 
    df = read_all_scenario_sheets(
        workbook_path=workbook_path,
        emission_cols=emission_cols,
        energy_cols=energy_cols,
        reference_year=reference_year,
    )
 
    path = (
        Path(output_dir)
        / f"{plant_name}_scenario_values.parquet"
    )
 
    df.write_parquet(path)
 
    return str(path)
 
 
def read_production_table(
    workbook_path: str,
    sheet_name: str = "Production",
) -> pl.DataFrame:
    # Read entire sheet without assuming a header
    raw = pl.read_excel(
        workbook_path,
        sheet_name=sheet_name,
        has_header=False,
    )
 
    # edge case - no capacity / flh / efficiency table
    if len(raw) <= 1:
        return pl.DataFrame()
 
 
    # Find row containing the table header
    header_row = (
        raw.with_row_index()
        .filter(
            (pl.col("column_1") == "Scenario")
            & (pl.col("column_2") == "Year")
        )
        .select("index")
        .item()
    )
 
    # Extract header values
    header_values = raw.row(header_row)
 
    # Keep columns until first null header
    n_cols = next(
        (i for i, v in enumerate(header_values) if v is None),
        len(header_values)
    )
 
    headers = [str(v) for v in header_values[:n_cols]]
 
    # Data below header
    data = raw.slice(header_row + 1)
 
    # Keep only relevant columns
    data = data.select(data.columns[:n_cols])
 
    # Rename columns
    data.columns = headers
 
    # Stop at first completely empty row
    empty_mask = pl.all_horizontal(
        [pl.col(c).is_null() for c in data.columns]
    )
 
    empty_rows = (
        data.with_row_index()
        .filter(empty_mask)
        .select("index")
        .to_series()
        .to_list()
    )
 
    if empty_rows:
        data = data.slice(0, empty_rows[0])
 
    return data
 
 
def save_production_values(
    workbook_path: str,
    plant_name: str,
    output_dir: str,
) -> str:
 
    df = read_production_table(
        workbook_path,
        sheet_name="Production",
    )
 
    path = (
        Path(output_dir)
        / f"{plant_name}_flh_production.parquet"
    )
 
    df.write_parquet(path)
 
    return str(path)
 

def read_production_table_curves(
        workbook_path:str=None,
        curves_df: pl.DataFrame = None,
        sheet_name:str = None
        ) -> pl.DataFrame:

    if workbook_path is not None:
        df =  pl.read_excel(workbook_path, sheet_name=sheet_name)
    else:
        df = curves_df
    
    df = df.rename(str.capitalize)
    df = df.rename({'Flh':'FLH'})
    df = df.with_columns(pl.col("Year").cast(pl.String))
    df = df.with_columns(pl.col('Cluster').str.to_lowercase().replace('overig', 'Cluster 6').alias('Cluster'))

    return df


def read_plant_details(workbook_path: str) -> dict:
    """
    Read plant details sheet and extract latitude, longitude.
    
    Returns:
        {"Latitude": float, "Longitude": float}
    """
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Plant details"]  # or whatever the sheet is called
    
    details = {}
    
    # Find the cells with Latitude and Longitude (scan first few rows)
    for row_idx in range(1, 20):
        for col_idx in range(1, 10):
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val == "Breedtegraad":
                value = ws.cell(row_idx, col_idx + 1).value
                details["Latitude"] = "" if value is None else value
            elif cell_val == "Lengtegraad":
                value = ws.cell(row_idx, col_idx + 1).value
                details["Longitude"] = "" if value is None else value
    
    return details
